#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FC-1 스킬 4단계: 조정 전표 수기 반영 + 열 그룹화.

3단계(add_allocation_basis.py) 결과 edit 파일에 대해:

  A) 기준정보 파일의 '1차 편집' 시트에서 '[조정 전표 추가]' 블록을 모두 찾아 edit 시트 맨 끝에
     행으로 추가한다. 이 시트가 아예 없거나 블록이 하나도 없으면(이번 달은 수기 조정이 없는 경우)
     이 부분은 건너뛰고 B)만 수행한다.
  B) edit 시트의 A~B, D~G, L, P, S~Y, AB~AD, AF~AG 열을 그룹화(outline)한다. 이건 조정 전표
     유무와 무관하게 매달 항상 적용한다.

## '1차 편집' 시트 구조

여러 조정 건이 번호 섹션(예: "1) IT 서비스 비용 조정")으로 나뉘어 있고, 각 섹션은 보통
  1. 섹션 제목 행
  2. (설명 문구, 있을 수도 없을 수도 있음)
  3. "원전표" 참고용 표: 헤더 행(본사/법인 ~ 배부기준) + 데이터 행 1개 이상
     (이 표는 edit 시트에 이미 존재하는 실제 원본 전표를 보여주는 참고용이라, 그대로 결합하면 안 된다)
  4. '[조정 전표 추가]' 마커 셀 (여기에 "및 상기 원전표의 배부기준도 'X'로 수정" 같은 추가 지시가
     같이 적혀 있을 수 있다)
  5. 조정 전표 표: 헤더 행(원전표 표와 동일한 헤더) + 실제로 edit 시트에 추가해야 하는 데이터 행들

헤더/마커 셀은 보통 B열(2번째 열)부터 시작하고 A열은 비어 있다 — 고정 열 번호에 의존하지 않고
'[조정 전표 추가]'라는 문자열이 들어있는 셀을 찾아서, 그 바로 아래 행을 헤더 행으로 삼는다.

## 조정 전표 데이터 행의 특수 지시문 처리

조정 전표 표의 셀 값이 아래처럼 실제 값이 아니라 "지시문"으로 들어있는 경우가 있다. 실제 값으로
치환해야 한다:

- Posting Date 열에 "당월 말일자로 기재" -> edit 시트 Posting Date의 최빈 연/월 기준 해당 월 말일.
- Posting Date/Document Number 열에 "상기 원전표와 동일 기재"(또는 "상기 원전표 날짜와 동일 기재"
  등 유사 문구) -> 바로 위 "원전표" 참고 표에 있는 값과 Cost Center/G/L Account/Vendor Name/
  금액(절대값)이 일치하는 edit 시트의 실제 행을 찾아 그 행의 Posting Date/Document Number를
  그대로 가져온다. 후보가 정확히 하나면 사용하고, 0개나 2개 이상이면 자동으로 채우지 않고 콘솔에
  경고를 남겨 사용자 확인을 받는다 (잘못 짝지어 엉뚱한 날짜를 넣는 것보다 안전).
- Currency 열에 "당월환율기재" -> 환율 폴더에서 해당 월 평균환율(Document currency 기준)을
  가져와 채운다 (1단계 build_edit_file.py의 환율 조회 로직과 동일한 방식).
- Amount(KRW) 열에 "당월환율적용 금액 기재" -> 값을 직접 넣지 않고, 다른 데이터 행과 동일하게
  `=Amount in local currency * Currency` 수식으로 채운다 (기존 Z열 패턴과 동일하게 유지해야
  필터/합계 수식이 깨지지 않는다).

## 원전표(기존 행) 자체를 수정해야 하는 경우

'[조정 전표 추가]' 마커 셀 텍스트에 "및 상기 원전표의 배부기준도 'X'로 수정" 같은 문구가 붙어
있으면, 위에서 "상기 원전표와 동일 기재"로 짝지은 원전표 행(edit 시트의 기존 행)의 배부기준 값도
X로 함께 수정한다. 원전표를 못 찾았으면 이 수정도 건너뛰고 경고한다.

## Vendor Name 공란 처리

조정 전표 Text가 감가상각비/상각비 성격(소계정에 "감가상각비" 또는 "상각비" 포함, 예: 사용권자산
감가상각비 대체 분개)이면 Vendor Name은 공란으로 둔다 — 기준정보 표에는 참고용으로 원전표의
Vendor Name이 같이 적혀 있는 경우가 있지만, 감가상각비 분개 자체는 특정 거래처에 대한 지급이
아니므로 회계상 거래처를 비워두는 게 맞다.

## 타입/정렬 일관성

Company Code/Cost Center/G/L Account 세 열은 (1단계에서와 동일하게) 항상 텍스트로 통일해서
넣는다 — 기존 데이터와 셀 타입이 섞이면 Excel에서 좌/우 정렬이 달라져 시각적으로 어긋나 보인다.
"""

import argparse
import glob
import os
import re
from copy import copy
from datetime import datetime, timedelta

import openpyxl
from openpyxl.utils import get_column_letter

MARKER = "[조정 전표 추가]"
CODE_HEADERS = ("Company Code", "Cost Center", "G/L Account")  # 항상 텍스트로 통일할 열
PER_100_CURRENCIES = {"JPY", "VND"}  # 환율추이 시트에 100단위로 기재되는 통화 -> 100으로 나눠서 사용


# ---------------------------------------------------------------------------
# 환율 (1단계 build_edit_file.py와 동일한 로직)
# ---------------------------------------------------------------------------

def find_fx_file(fx_folder):
    candidates = glob.glob(os.path.join(fx_folder, "환율 추이*.xlsx")) + \
                 glob.glob(os.path.join(fx_folder, "환율*.xlsx"))
    candidates = [c for c in candidates if not os.path.basename(c).startswith("~$")]
    if not candidates:
        raise FileNotFoundError(f"환율 폴더({fx_folder})에서 '환율 추이*.xlsx' 파일을 찾지 못했습니다.")
    return max(candidates, key=os.path.getmtime)


def load_fx_rates(fx_path, yy, mm):
    wb = openpyxl.load_workbook(fx_path, data_only=True)
    if "환율추이" not in wb.sheetnames:
        raise ValueError(f"'{fx_path}'에 '환율추이' 시트가 없습니다.")
    ws = wb["환율추이"]
    target_labels = {f"{yy}.{mm:02d}", f"'{yy}.{mm:02d}"}
    header_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=3, column=c).value
        if v is not None and str(v).strip() in target_labels:
            header_col = c
            break
    if header_col is None:
        raise ValueError(f"환율 파일에서 {yy}.{mm:02d} 열을 찾지 못했습니다.")
    avg_col = header_col + 1
    if ws.cell(row=4, column=avg_col).value != "평균":
        for c in range(header_col, header_col + 3):
            if ws.cell(row=4, column=c).value == "평균":
                avg_col = c
                break
    rates = {"KRW": 1}
    for r in range(5, ws.max_row + 1):
        cur = ws.cell(row=r, column=1).value
        if not cur:
            continue
        cur_str = str(cur).strip()
        val = ws.cell(row=r, column=avg_col).value
        if val is not None:
            # JPY, VND는 100단위 기준으로 기재되어 있으므로 100으로 나눠서 1단위 환율로 환산한다
            if cur_str in PER_100_CURRENCIES:
                val = val / 100
            rates[cur_str] = val
    return rates


def detect_month_from_edit(ws, post_date_col):
    from collections import Counter
    counter = Counter()
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=post_date_col).value
        if hasattr(v, "year"):
            counter[(v.year, v.month)] += 1
    if not counter:
        raise ValueError("edit 시트의 Posting Date에서 대상 월을 확인할 수 없습니다.")
    (yy, mm), _ = counter.most_common(1)[0]
    return yy, mm


def month_last_day(year, month):
    if month == 12:
        nxt = datetime(year + 1, 1, 1)
    else:
        nxt = datetime(year, month + 1, 1)
    return nxt - timedelta(days=1)


# ---------------------------------------------------------------------------
# 값 정규화
# ---------------------------------------------------------------------------

def to_text(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def norm(v):
    t = to_text(v)
    return t


# ---------------------------------------------------------------------------
# '1차 편집' 시트 파싱
# ---------------------------------------------------------------------------

def find_marker_cells(ws):
    """'[조정 전표 추가]'로 시작하는 셀을 모두 찾아 (row, col, 전체텍스트) 리스트로 반환."""
    out = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().startswith(MARKER):
                out.append((r, c, v.strip()))
    return out


def read_table(ws, header_row, col_start, max_span=45, blank_tolerance=2):
    """header_row를 헤더로 삼아 col_start부터 헤더 이름을 읽는다. edit 시트와 동일하게 URL과
    배부기준 사이에 빈 칸이 한 칸 끼어 있으므로, 헤더가 없는 열이 나와도 바로 멈추지 않고
    `blank_tolerance`번까지는 건너뛰고 계속 읽는다 (그 이상 연속으로 비어 있으면 표가 끝난 것으로
    본다). 헤더는 {헤더명: 열번호} 형태로 저장해서 열 위치가 밀려 있어도 이름으로 찾을 수 있다."""
    col_of = {}
    consec_blank = 0
    c = col_start
    while c < col_start + max_span:
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            consec_blank += 1
            if consec_blank > blank_tolerance:
                break
        else:
            consec_blank = 0
            col_of[v] = c
        c += 1
    headers = list(col_of.keys())
    rows = []
    r = header_row + 1
    while True:
        vals = {h: ws.cell(row=r, column=c2).value for h, c2 in col_of.items()}
        if all(v is None for v in vals.values()):
            break
        rows.append(vals)
        r += 1
    return headers, rows, r  # r = 다음 빈 행 (호출부에서 다음 섹션 탐색 시작점으로 사용 가능)


def find_reference_table_above(ws, marker_row, marker_col):
    """마커 바로 위에 있는 '원전표' 참고 표(헤더 행 + 데이터 행들)를 찾는다.
    마커 위로 올라가면서 이 열(marker_col)에 '본사/법인' 헤더가 나오는 첫 행을 헤더 행으로 삼는다."""
    r = marker_row - 1
    while r > 0:
        v = ws.cell(row=r, column=marker_col).value
        if v == "본사/법인":
            return read_table(ws, r, marker_col)
        r -= 1
    return None, [], None


# ---------------------------------------------------------------------------
# edit 시트에서 원전표(실제 행) 매칭
# ---------------------------------------------------------------------------

def find_matching_edit_row(ws, idx, ref_row):
    """참고 표의 원전표 한 행(ref_row dict)과 Cost Center/G/L Account/Vendor Name/금액(절대값)이
    일치하는 edit 시트의 실제 행을 찾는다. 정확히 하나면 그 행 번호를, 아니면 None(+후보 목록)을
    반환한다."""
    gl = norm(ref_row.get("G/L Account"))
    cc = norm(ref_row.get("Cost Center"))
    vendor = norm(ref_row.get("Vendor Name"))
    amt = ref_row.get("Amount in local currency")

    gl_idx = idx.get("G/L Account")
    cc_idx = idx.get("Cost Center")
    vendor_idx = idx.get("Vendor Name")
    amt_idx = idx.get("Amount in local currency")
    text_idx = idx.get("Text")

    candidates = []
    for r in range(3, ws.max_row + 1):
        if gl_idx and norm(ws.cell(row=r, column=gl_idx).value) != gl:
            continue
        if cc and cc_idx and norm(ws.cell(row=r, column=cc_idx).value) != cc:
            continue
        if vendor and vendor_idx and norm(ws.cell(row=r, column=vendor_idx).value) != vendor:
            continue
        if amt is not None and amt_idx:
            row_amt = ws.cell(row=r, column=amt_idx).value
            if not isinstance(row_amt, (int, float)) or abs(abs(row_amt) - abs(amt)) > 0.01:
                continue
        # 이전에 이 스크립트가 추가한 [Manual] 행은 원전표 후보에서 제외
        if text_idx:
            t = ws.cell(row=r, column=text_idx).value
            if isinstance(t, str) and t.startswith("[Manual]"):
                continue
        candidates.append(r)

    if len(candidates) == 1:
        return candidates[0], candidates
    return None, candidates


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------

GROUP_RANGES = [("A", "B"), ("D", "G"), ("L", "L"), ("P", "P"), ("S", "Y"), ("AB", "AD"), ("AF", "AG")]
DEPRECIATION_HINTS = ("감가상각비", "상각비")


def build(edit_path, ref_path, fx_folder, output_path):
    wb = openpyxl.load_workbook(edit_path, data_only=False)
    ws = wb["edit"]
    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    for col in ("Posting Date", "Document Number", "Cost Center", "G/L Account", "Currency",
                "Amount in local currency", "Text", "Vendor Name", "배부기준", "소계정"):
        if col not in idx:
            raise ValueError(f"edit 시트에 '{col}' 열이 없습니다. 3단계(add_allocation_basis.py)를 먼저 실행하세요.")

    post_date_idx = idx["Posting Date"]
    docnum_idx = idx["Document Number"]
    currency_idx = idx["Currency"]
    amt_local_idx = idx["Amount in local currency"]
    amt_krw_idx = idx["Amount(KRW)"]
    doc_currency_idx = idx["Document currency"]
    text_idx = idx["Text"]
    vendor_idx = idx["Vendor Name"]
    basis_idx = idx["배부기준"]
    subj_idx = idx["소계정"]

    amt_local_letter = get_column_letter(amt_local_idx)
    currency_letter = get_column_letter(currency_idx)

    yy2, mm = detect_month_from_edit(ws, post_date_idx)
    yy = yy2 % 100
    month_end = month_last_day(yy2, mm)
    fx_path = None
    fx_rates = None

    ref_wb = openpyxl.load_workbook(ref_path, data_only=True)
    added_rows = 0
    review_needed = []
    original_row_updates = []

    if "1차 편집" not in ref_wb.sheetnames:
        print("기준정보 파일에 '1차 편집' 시트가 없습니다 — 이번 달은 수기 조정 전표가 없는 것으로 보고 건너뜁니다.")
    else:
        edit_sheet = ref_wb["1차 편집"]
        markers = find_marker_cells(edit_sheet)
        print(f"'[조정 전표 추가]' 블록 {len(markers)}건 발견")

        template_row = ws.max_row  # 스타일 복사용 (기존 마지막 데이터 행)

        for marker_row, marker_col, marker_text in markers:
            # 배부기준 재기재 지시문 파싱: "배부기준...'X'로 수정"
            m = re.search(r"배부\s*기준.*?'([^']+)'\s*로\s*수정", marker_text)
            basis_override = m.group(1) if m else None

            ref_headers, ref_rows, _ = find_reference_table_above(edit_sheet, marker_row, marker_col)
            adj_headers, adj_rows, _ = read_table(edit_sheet, marker_row + 1, marker_col)

            if not adj_rows:
                print(f"  [경고] {marker_row}행 블록에 조정 전표 데이터 행이 없습니다 — 건너뜀")
                continue

            # 참고 표(원전표)는 한 블록 안에 여러 건이 들어있을 수 있다 (예: 임차료 블록에 HR실
            # 건과 SCM실 건이 같이 있는 경우). 각 원전표 후보를 미리 개별적으로 매칭해두고,
            # 조정 전표 행마다 자신의 Cost Center(없으면 G/L Account+Company)로 어느 원전표에
            # 대응하는지 골라 쓴다 — 블록 전체에 하나의 원전표만 있다고 가정하면 안 된다.
            ref_matches = []
            for ref_row in ref_rows:
                r, candidates = find_matching_edit_row(ws, idx, ref_row)
                ref_matches.append((ref_row, r))
                if r is None:
                    review_needed.append(
                        f"{marker_row}행 블록: 원전표 후보를 자동으로 특정하지 못했습니다 "
                        f"(참고행 CC={ref_row.get('Cost Center')!r}, 후보 {len(candidates)}건) — "
                        f"해당 원전표에 의존하는 날짜/전표번호 지시문과 배부기준 재기재를 건너뜁니다."
                    )

            def pick_original_row(adj_row):
                if len(ref_matches) == 1:
                    return ref_matches[0][1]
                adj_cc = norm(adj_row.get("Cost Center"))
                if adj_cc:
                    for ref_row, r in ref_matches:
                        if norm(ref_row.get("Cost Center")) == adj_cc:
                            return r
                adj_gl = norm(adj_row.get("G/L Account"))
                adj_company = norm(adj_row.get("Company"))
                for ref_row, r in ref_matches:
                    if norm(ref_row.get("G/L Account")) == adj_gl and norm(ref_row.get("Company")) == adj_company:
                        return r
                return None

            if basis_override:
                for ref_row, r in ref_matches:
                    if r is not None:
                        original_row_updates.append((r, basis_override))
                    else:
                        review_needed.append(
                            f"{marker_row}행 블록: 원전표를 못 찾아 배부기준 재기재('{basis_override}')를 "
                            f"건너뜁니다 (참고행 CC={ref_row.get('Cost Center')!r}) — 직접 확인 필요."
                        )

            for adj in adj_rows:
                new_row = ws.max_row + 1
                matched_original_row = pick_original_row(adj)

                post_date_val = adj.get("Posting Date")
                docnum_val = adj.get("Document Number")

                if isinstance(post_date_val, str) and "말일" in post_date_val:
                    post_date_val = month_end
                elif isinstance(post_date_val, str) and "원전표" in post_date_val:
                    if matched_original_row is not None:
                        post_date_val = ws.cell(row=matched_original_row, column=post_date_idx).value
                    else:
                        post_date_val = None
                        review_needed.append(
                            f"{new_row}행(신규): 원전표를 특정하지 못해 Posting Date를 비워뒀습니다 — 직접 확인 필요."
                        )

                if isinstance(docnum_val, str) and "원전표" in docnum_val:
                    if matched_original_row is not None:
                        docnum_val = ws.cell(row=matched_original_row, column=docnum_idx).value
                    else:
                        docnum_val = None
                        review_needed.append(
                            f"{new_row}행(신규): 원전표를 특정하지 못해 Document Number를 비워뒀습니다 — 직접 확인 필요."
                        )

                for h, v in adj.items():
                    if h not in idx:
                        continue
                    col = idx[h]

                    if h == "Currency" and isinstance(v, str) and "환율" in v:
                        if fx_rates is None:
                            fx_path = find_fx_file(fx_folder)
                            fx_rates = load_fx_rates(fx_path, yy, mm)
                            print(f"  환율 파일: {os.path.basename(fx_path)} ({yy2}.{mm:02d} 평균환율 사용)")
                        doc_cur = adj.get("Document currency")
                        rate = fx_rates.get(doc_cur)
                        if rate is None:
                            review_needed.append(
                                f"{marker_row}행 블록: 통화 '{doc_cur}'의 환율을 찾지 못했습니다 — Currency 열 확인 필요."
                            )
                        v = rate
                    elif h == "Amount(KRW)" and isinstance(v, str) and "환율" in v:
                        continue  # 아래에서 별도로 수식으로 채움
                    elif h == "Posting Date":
                        v = post_date_val
                    elif h == "Document Number":
                        v = docnum_val
                    elif h in CODE_HEADERS:
                        v = to_text(v)

                    cell = ws.cell(row=new_row, column=col)
                    cell.value = v

                # Amount(KRW)은 항상 수식 (=Amount in local currency * Currency), 기존 패턴과 동일
                ws.cell(row=new_row, column=amt_krw_idx).value = (
                    f"={amt_local_letter}{new_row}*{currency_letter}{new_row}"
                )

                # 감가상각비/상각비 성격 조정은 Vendor Name 공란 처리
                subj_val = adj.get("소계정") or ws.cell(row=new_row, column=subj_idx).value
                if isinstance(subj_val, str) and any(h in subj_val for h in DEPRECIATION_HINTS):
                    ws.cell(row=new_row, column=vendor_idx).value = None

                # 스타일(글꼴/테두리/채우기/표시형식/정렬)을 기존 마지막 데이터 행에서 복사
                for col in range(1, ws.max_column + 1):
                    src = ws.cell(row=template_row, column=col)
                    dst = ws.cell(row=new_row, column=col)
                    dst.font = copy(src.font)
                    dst.border = copy(src.border)
                    dst.fill = copy(src.fill)
                    dst.number_format = src.number_format
                    dst.alignment = copy(src.alignment)
                ws.cell(row=new_row, column=post_date_idx).number_format = ws.cell(
                    row=template_row, column=post_date_idx).number_format
                ws.row_dimensions[new_row].height = ws.row_dimensions[template_row].height

                added_rows += 1

        for r, basis_val in original_row_updates:
            ws.cell(row=r, column=basis_idx).value = basis_val
            print(f"  원전표 {r}행 배부기준 -> '{basis_val}'로 수정")

    # 배부기준(원전표) 재기재는 값 자체가 텍스트라 타입 이슈 없음.
    # 새로 추가한 행들의 Company Code/Cost Center/G/L Account 타입은 위에서 이미 텍스트로 통일했지만,
    # 혹시 기존 raw 데이터 쪽에 아직 남아있는 숫자 타입도 이번 기회에 전체 통일한다.
    converted = 0
    for h in CODE_HEADERS:
        col = idx[h]
        for r in range(3, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None and not isinstance(v, str):
                ws.cell(row=r, column=col).value = to_text(v)
                converted += 1
    if converted:
        print(f"기존 데이터 중 {h} 등 코드성 열의 숫자 타입 {converted}개 셀을 텍스트로 통일 (정렬 일관성 확보)")

    # SUBTOTAL(V1/X1/Z1) 범위를 새 마지막 행까지 확장
    last_row = ws.max_row
    amt_doc_idx = idx["Amount in doc. curr."]
    for col in (amt_doc_idx, amt_local_idx, amt_krw_idx):
        letter = get_column_letter(col)
        cell = ws.cell(row=1, column=col)
        if isinstance(cell.value, str) and cell.value.startswith("=SUBTOTAL"):
            cell.value = f"=SUBTOTAL(109,{letter}3:{letter}{last_row})"

    # Z열 수식이 전체 행에서 일관되게 =X{row}*Y{row} 패턴인지 마지막으로 한 번 더 보정
    for r in range(3, last_row + 1):
        expected = f"={amt_local_letter}{r}*{currency_letter}{r}"
        if ws.cell(row=r, column=amt_krw_idx).value != expected:
            ws.cell(row=r, column=amt_krw_idx).value = expected

    # 배부기준 = '실적 제외'인 행은 Amount(KRW) 금액을 삭제 (매달 항상 적용)
    # 위의 Z열 수식 일괄 보정 다음에 실행해야 이 삭제가 덮어써지지 않는다.
    excluded_cleared = 0
    for r in range(3, last_row + 1):
        if ws.cell(row=r, column=basis_idx).value == "실적 제외":
            cell = ws.cell(row=r, column=amt_krw_idx)
            if cell.value is not None:
                cell.value = None
                excluded_cleared += 1

    # 열 그룹화 (매달 항상 적용)
    for start, end in GROUP_RANGES:
        ws.column_dimensions.group(start, end, outline_level=1, hidden=False)

    wb.save(output_path)

    print(f"완료: {output_path}")
    print(f"수기 조정 전표 추가: {added_rows}행 (마지막 행 -> {last_row})")
    print(f"원전표 배부기준 수정: {len(original_row_updates)}건")
    print(f"배부기준='실적 제외' -> Amount(KRW) 삭제: {excluded_cleared}건")
    print(f"열 그룹화 적용: {GROUP_RANGES}")
    if review_needed:
        print("[사용자 확인 필요]")
        for msg in review_needed:
            print(f"  - {msg}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--edit", required=True, help="3단계(add_allocation_basis.py) 결과 edit 파일 경로")
    ap.add_argument("--ref", required=True, help="기준정보 파일 경로 ('1차 편집' 시트 포함)")
    ap.add_argument("--fx-folder", required=True, help="환율 파일들이 있는 폴더 경로")
    ap.add_argument("--output", required=True)
    args = ap.parse_args()
    build(args.edit, args.ref, args.fx_folder, args.output)
