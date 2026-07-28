"""
FC-1 스킬 3단계: 배부기준 열 추가.

2단계(add_reference_cols.py) 결과 edit 파일에 대해:
  1) 테이블 끝(원래 33열, URL)에서 한 열 띄고 '배부기준' 열을 새로 만든다 (헤더는 34번째 열은 비우고
     35번째 열에 작성). 이미 '배부기준' 열이 있으면(재작업 시) 그 열을 그대로 재사용해 값만 다시 채운다.
  2) 각 데이터 행에 대해 아래 순서로 배부기준 값을 결정한다.
     a. 기본값: 배부조직 열 값을 그대로 복사.
     b. 기준정보 파일의 '배부기준' 시트 표(Cost Center / 대조직 / 배부조직 / G/L Account / 대계정(re)
        중 해당 행에 값이 채워진 열들을 AND 조건으로 사용)를 위에서 아래 순서로 순차 확인하며, 조건이
        일치할 때마다 배부기준 값을 그 행의 결과 값으로 덮어쓴다 (나중에 매칭되는 조건이 이전 매칭을
        덮어쓴다 - 즉 마지막으로 매칭된 조건이 최종값이 된다). 표는 일반적인 규칙을 위쪽에, 특정
        Cost Center/GL 조합처럼 더 구체적인 예외 규칙을 아래쪽에 배치하는 방식으로 관리되므로, 이
        "나중 매칭 우선" 방식이어야 예외 규칙이 일반 규칙을 올바르게 덮어쓴다.
     c. [항상 가장 마지막에 적용] 보고용 열 값이 '7. 해외법인(법인)'이면 배부기준 값을 Company 열
        값으로 최종 덮어쓴다 (표 조건 매칭 여부와 무관하게 이 규칙이 최우선/최종 적용된다).
  3) 보고용 열 값에 '주재원'이 포함되어 있으면(예: '6.해외법인(주재원)') 본사/법인 열 값을 '법인'으로
     재기재한다 (원래 raw 시트명이 본사_ 접두사라 '본사'로 채워져 있던 것을 교정).

기준정보 파일의 '배부기준' 시트 구조:
  - 상단에 안내 문구 몇 줄이 있고, 그 아래 'Cost Center' 로 시작하는 헤더 행이 나온다. 헤더 행 위치는
    문서용 안내 줄 수가 달라질 수 있으므로 고정 행 번호를 쓰지 않고 A열 값이 'Cost Center'인 행을
    찾아서 사용한다.
  - 헤더: Cost Center, Cost Ctr Name, 대조직, 배부조직, G/L Account, 소계정, 대계정(re), 배부기준, 비고
  - 매칭에는 Cost Center / 대조직 / 배부조직 / G/L Account / 대계정(re) 5개 열만 사용한다.
    Cost Ctr Name과 소계정은 각각 Cost Center/G/L Account의 설명용 열이라 매칭 조건에서 제외한다
    (값이 raw와 다르게 정리되어 있거나 공백 등 사소한 차이가 있을 수 있어 오매칭 위험을 줄이기 위함).
  - 한 행에 여러 열이 채워져 있으면 전부 AND 조건이다 (예: Cost Center=105953 이면서 G/L Account=537600).
"""

import argparse
import openpyxl
from copy import copy


def norm(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def find_header_row(ws):
    for r in range(1, 20):
        if ws.cell(row=r, column=1).value == 'Cost Center':
            return r
    raise ValueError("배부기준 시트에서 헤더 행을 찾지 못했습니다 (A열에 'Cost Center' 없음). "
                      "기준정보 파일의 '배부기준' 시트 구조가 바뀐 것일 수 있습니다.")


def load_rules(ref_path):
    wb = openpyxl.load_workbook(ref_path, data_only=True)
    if '배부기준' not in wb.sheetnames:
        raise ValueError("기준정보 파일에 '배부기준' 시트가 없습니다.")
    ws = wb['배부기준']
    header_row = find_header_row(ws)
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, 10)]
    assert (headers[0] == 'Cost Center' and headers[2] == '대조직' and headers[3] == '배부조직'
            and headers[4] == 'G/L Account' and headers[6] == '대계정(re)' and headers[7] == '배부기준'), \
        f"배부기준 시트 헤더가 예상과 다릅니다: {headers}"

    rules = []
    for r in range(header_row + 1, ws.max_row + 1):
        result = ws.cell(row=r, column=8).value
        if result is None:
            continue
        cc = norm(ws.cell(row=r, column=1).value)
        dj = norm(ws.cell(row=r, column=3).value)
        bj = norm(ws.cell(row=r, column=4).value)
        gl = norm(ws.cell(row=r, column=5).value)
        dgre = norm(ws.cell(row=r, column=7).value)
        conds = {}
        if cc is not None:
            conds['cc'] = cc
        if dj is not None:
            conds['dj'] = dj
        if bj is not None:
            conds['bj'] = bj
        if gl is not None:
            conds['gl'] = gl
        if dgre is not None:
            conds['dgre'] = dgre
        if not conds:
            continue
        rules.append((r, conds, result))
    return rules, header_row


def build(edit_path, ref_path, output_path):
    rules, header_row = load_rules(ref_path)
    print(f"[배부기준 시트 헤더행={header_row}] 조건 규칙 {len(rules)}건 로드")

    wb = openpyxl.load_workbook(edit_path)
    ws = wb['edit']
    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    for col in ('Company Code', 'Cost Center', 'G/L Account', '대조직', '배부조직', '대계정(re)', '보고용', 'Company', '본사/법인'):
        if col not in idx:
            raise ValueError(f"edit 시트에 '{col}' 열이 없습니다. 2단계(add_reference_cols.py)를 먼저 실행하세요.")

    cocd_idx = idx['Company Code']
    cc_idx = idx['Cost Center']
    gl_idx = idx['G/L Account']
    dj_idx = idx['대조직']
    bj_idx = idx['배부조직']
    dgre_idx = idx['대계정(re)']
    report_idx = idx['보고용']
    company_idx = idx['Company']
    bonsa_idx = idx['본사/법인']
    sample_col = idx['Company']  # 새 열 스타일(진한 회색 헤더)을 복사해올 기존 "새 컬럼" 샘플

    # '배부기준' 열이 이미 있으면(재작업) 재사용, 없으면 테이블 끝에서 한 칸 띄우고 새로 생성
    if '배부기준' in idx:
        basis_col = idx['배부기준']
        print(f"기존 '배부기준' 열(컬럼 {basis_col}) 재사용")
    else:
        last_col = ws.max_column
        basis_col = last_col + 2  # last_col+1 은 빈 칸으로 남긴다
        sample_header_cell = ws.cell(row=2, column=sample_col)
        header_cell = ws.cell(row=2, column=basis_col)
        header_cell.value = '배부기준'
        header_cell.fill = copy(sample_header_cell.fill)
        header_cell.font = copy(sample_header_cell.font)
        header_cell.border = copy(sample_header_cell.border)
        header_cell.alignment = copy(sample_header_cell.alignment)
        print(f"'배부기준' 열 신규 생성 (컬럼 {basis_col}, 빈 칸은 컬럼 {last_col + 1})")

    overseas_override_count = 0
    e100_override_count = 0
    e100_depre_count = 0
    jujaewon_count = 0
    rule_match_counts = {r[0]: 0 for r in rules}

    for r in range(3, ws.max_row + 1):
        cocd = norm(ws.cell(row=r, column=cocd_idx).value)
        cc = norm(ws.cell(row=r, column=cc_idx).value)
        gl = norm(ws.cell(row=r, column=gl_idx).value)
        dj = norm(ws.cell(row=r, column=dj_idx).value)
        bj = norm(ws.cell(row=r, column=bj_idx).value)
        dgre = norm(ws.cell(row=r, column=dgre_idx).value)
        report = ws.cell(row=r, column=report_idx).value
        company = ws.cell(row=r, column=company_idx).value

        row_vals = {'cc': cc, 'dj': dj, 'bj': bj, 'gl': gl, 'dgre': dgre}

        # 1) 기본값: 배부조직 그대로 복사
        value = bj

        # 2) 배부기준 표 조건을 위->아래 순서로 순차 적용 (나중 매칭이 이전 매칭을 덮어씀)
        for rownum, conds, result in rules:
            if all(row_vals.get(k) == v for k, v in conds.items()):
                value = result
                rule_match_counts[rownum] += 1

        # 3) [항상 마지막에 적용, 배부기준 시트 안내에 명시된 순서]
        #    a. 보고용이 '7. 해외법인(법인)'이면 Company 값으로 최종 덮어쓰기
        if report == '7. 해외법인(법인)':
            value = company
            overseas_override_count += 1

        #    b. Company Code가 'E100'이면 'EVCS(국내) 100%'로, 대계정(re)가 '23 감가상각비'이면
        #       'EVCS(국내/해외)'로 최종 덮어쓰기 (배부기준 시트 하단 "3)" 안내에 명시된 신규 규칙)
        if cocd == 'E100':
            if dgre == '23 감가상각비':
                value = 'EVCS(국내/해외)'
                e100_depre_count += 1
            else:
                value = 'EVCS(국내) 100%'
            e100_override_count += 1

        cell = ws.cell(row=r, column=basis_col)
        cell.value = value
        sample_data_cell = ws.cell(row=r, column=sample_col)
        cell.font = copy(sample_data_cell.font)
        cell.alignment = copy(sample_data_cell.alignment)

        # 4) 보고용에 '주재원'이 포함되면 본사/법인 열을 '법인'으로 재기재
        if report is not None and '주재원' in str(report):
            ws.cell(row=r, column=bonsa_idx).value = '법인'
            jujaewon_count += 1

    wb.save(output_path)

    print(f"완료: {output_path}")
    print(f"총 데이터 행: {ws.max_row - 2}")
    print(f"'7. 해외법인(법인)' -> Company 값 최종 대체 건수: {overseas_override_count}")
    print(f"Company Code='E100' -> 배부기준 최종 대체 건수: {e100_override_count} (그 중 '23 감가상각비' -> 'EVCS(국내/해외)': {e100_depre_count}건, 나머지는 'EVCS(국내) 100%')")
    print(f"'주재원' -> 본사/법인='법인' 변경 건수: {jujaewon_count}")
    print("표 조건 규칙별 매칭 건수 (0건 제외):")
    for rownum, conds, result in rules:
        c = rule_match_counts[rownum]
        if c > 0:
            print(f"  row{rownum} {conds} -> '{result}' : {c}건")


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--edit', required=True, help='2단계(add_reference_cols.py) 결과 edit 파일 경로')
    ap.add_argument('--ref', required=True, help="기준정보 파일 경로 ('배부기준' 시트 포함)")
    ap.add_argument('--output', required=True)
    args = ap.parse_args()
    build(args.edit, args.ref, args.output)
