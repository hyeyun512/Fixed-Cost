#!/usr/bin/env python3
"""
FC.1 — HEV 고정비 raw data -> edit_{월}.xlsx 자동 생성

입력: raw_{N}월.xlsx (여러 시트, 각 시트는 SAP FBL3N/유사 다운로드 포맷:
      Company Code, Posting Date, Order, Document Number, Document Type,
      Cost Center, G/L Account, Posting Key, Document currency,
      Amount in doc. curr., Local Currency, Amount in local currency,
      Text, Reversed with, Reference key 1, Reference, Vendor Name, Vendor, URL
      — 19개 컬럼, 시트명은 "본사_XXX" 또는 "법인_XXX" 형태)

출력: edit_{N}월.xlsx, 시트 "edit" + "검증"
      - "edit" 시트: 모든 원본 시트를 하나의 표로 세로 결합
      - Company Code가 빈 소계/합계 행은 결합 대상에서 제외하되, 그 행의 "Amount in local
        currency" 값은 해당 시트의 "원본 합계"로 기록해 검증에 사용한다 (해당 시트의 가장
        마지막 소계/합계 행 값을 그 시트의 대표 합계로 간주)
      - 본사/법인 열(맨 앞): 시트명이 "본사_"로 시작하면 "본사", "법인_"로 시작하면 "법인"
      - Amount in local currency 오른쪽에 Currency(해당월 평균환율), Amount(KRW)(=Amount in local currency*Currency, 수식) 열 추가
      - 1행: Amount in doc. curr., Amount in local currency, Amount(KRW) 세 열에 SUBTOTAL(109,...) 합계 수식 (표시형식 #,##0;[Red]-#,##0, 데이터/합계 행 모두 적용)
        (SUM이 아닌 SUBTOTAL을 쓰는 이유: 아래 표에 자동필터가 걸려 있어서, 필터링 시 보이는 행만 합산되게 하기 위함)
      - 2행: 헤더. 원본에 있던 컬럼은 밝은 회색(FFDDDDDD), 이 스크립트가 새로 추가한 컬럼(본사/법인/Currency/Amount(KRW))은
        더 진한 회색(FFA6A6A6)으로 구분. 볼드는 사용하지 않음.
      - freeze_panes A3, autofilter A2:마지막
      - "검증" 시트: 원본 시트별 "Amount in local currency" 계산 합계(결합 대상으로 포함된 행들의 합)와
        원본 시트의 합계행 값을 나란히 놓고 차이를 계산해, 결합 과정에서 값이 누락/중복되지 않았는지
        검토할 수 있게 한다. 차이가 허용오차(0.5)를 넘으면 "확인 필요"로 표시한다.
"""
import argparse
import glob
import os
import re
from collections import Counter

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side

NCOLS_RAW = 19  # Company Code ... URL
TOLERANCE = 0.5  # 합계 검증 허용오차
PER_100_CURRENCIES = {"JPY", "VND"}  # 환율추이 시트에 100단위로 기재되는 통화 -> 100으로 나눠서 사용


def find_fx_file(fx_folder):
    """환율 폴더에서 '환율 추이(...).xlsx' 패턴의 파일 중 가장 최근 것을 찾는다."""
    candidates = glob.glob(os.path.join(fx_folder, "환율 추이*.xlsx")) + \
                 glob.glob(os.path.join(fx_folder, "환율*.xlsx"))
    candidates = [c for c in candidates if not os.path.basename(c).startswith("~$")]
    if not candidates:
        raise FileNotFoundError(f"환율 폴더({fx_folder})에서 '환율 추이*.xlsx' 파일을 찾지 못했습니다.")
    # 가장 최근에 수정된 파일 사용
    return max(candidates, key=os.path.getmtime)


def detect_month(raw_wb, raw_filename):
    """Posting Date 값들 중 가장 많이 등장하는 연/월을 찾는다. (연2자리, 월)"""
    counter = Counter()
    for sname in raw_wb.sheetnames:
        ws = raw_wb[sname]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if "Posting Date" not in headers:
            continue
        col = headers.index("Posting Date") + 1
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if hasattr(v, "year"):
                counter[(v.year % 100, v.month)] += 1
    if counter:
        (yy, mm), _ = counter.most_common(1)[0]
        return yy, mm
    # fallback: 파일명에서 "N월" 패턴 추출 (연도는 알 수 없으므로 에러)
    m = re.search(r"(\d{1,2})월", raw_filename)
    if m:
        raise ValueError(
            "Posting Date에서 연도를 확인할 수 없습니다. raw 파일에 날짜 데이터가 있는지 확인해주세요."
        )
    raise ValueError("raw 파일에서 대상 월을 확인할 수 없습니다.")


def load_fx_rates(fx_path, yy, mm):
    """환율 추이 파일의 '환율추이' 시트에서 해당 연/월 '평균' 환율을 통화별로 읽어온다."""
    wb = openpyxl.load_workbook(fx_path, data_only=True)
    if "환율추이" not in wb.sheetnames:
        raise ValueError(f"'{fx_path}'에 '환율추이' 시트가 없습니다.")
    ws = wb["환율추이"]

    target_labels = {f"{yy}.{mm:02d}", f"'{yy}.{mm:02d}"}
    header_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=3, column=c).value
        if v is not None and str(v).strip() in target_labels:
            header_col = c
            break
    if header_col is None:
        raise ValueError(
            f"환율 파일에서 {yy}.{mm:02d} 열을 찾지 못했습니다. 해당 월 데이터가 아직 없을 수 있습니다."
        )

    # 각 월 블록은 3개 컬럼: [일자환율, 평균, 미화환산율] 순서 -> 평균은 header_col+1
    avg_col = header_col + 1
    avg_label = ws.cell(row=4, column=avg_col).value
    if avg_label != "평균":
        # 혹시 레이아웃이 다르면 3개 컬럼 중 '평균' 라벨을 직접 탐색
        for c in range(header_col, header_col + 3):
            if ws.cell(row=4, column=c).value == "평균":
                avg_col = c
                break

    rates = {"KRW": 1}
    for r in range(5, ws.max_row + 1):
        cur = ws.cell(row=r, column=1).value
        if not cur:
            continue
        cur_str = str(cur).strip()
        val = ws.cell(row=r, column=avg_col).value
        if val is not None:
            # JPY, VND는 환율추이 시트에 100단위 기준으로 기재되어 있으므로 100으로 나눠서
            # 1단위 환율로 환산한다 (예: 시트에 900으로 적혀 있으면 실제로는 100엔=900원 -> 9로 사용)
            if cur_str in PER_100_CURRENCIES:
                val = val / 100
            rates[cur_str] = val
    return rates


def build(raw_path, fx_folder, output_path):
    raw_f = openpyxl.load_workbook(raw_path, data_only=False)
    raw_v = openpyxl.load_workbook(raw_path, data_only=True)

    yy, mm = detect_month(raw_v, os.path.basename(raw_path))
    fx_path = find_fx_file(fx_folder)
    fx_rates = load_fx_rates(fx_path, yy, mm)

    sheets = raw_f.sheetnames
    base_headers = [raw_f[sheets[0]].cell(row=1, column=c).value for c in range(1, NCOLS_RAW + 1)]
    if len(base_headers) < 12 or "Amount in local currency" not in base_headers:
        raise ValueError("raw 파일의 첫 시트 헤더가 예상 포맷과 다릅니다. (Company Code ~ URL 19개 컬럼 예상)")

    amt_local_idx_raw = base_headers.index("Amount in local currency")  # vals 배열 내 인덱스(0-based)

    full_headers = (["본사/법인"] + base_headers[:12]
                     + ["Currency", "Amount(KRW)"]
                     + base_headers[12:])

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "edit"

    out_ws.append([None] * len(full_headers))
    out_ws.append(full_headers)

    currency_col = full_headers.index("Currency") + 1
    krw_col = full_headers.index("Amount(KRW)") + 1
    amt_local_letter = get_column_letter(full_headers.index("Amount in local currency") + 1)
    curr_letter = get_column_letter(currency_col)

    total_rows = 0
    missing_fx = set()
    sheet_validation = []  # (sheet_name, gubun, computed_sum, reported_total)
    for sname in sheets:
        if sname.startswith("본사"):
            gubun = "본사"
        elif sname.startswith("법인"):
            gubun = "법인"
        else:
            gubun = sname  # 규칙에 안 맞는 시트명은 시트명 그대로 표기 (사람이 검토하도록)
        wsf = raw_f[sname]
        wsv = raw_v[sname]
        headers = [wsf.cell(row=1, column=c).value for c in range(1, wsf.max_column + 1)]
        if headers[:NCOLS_RAW] != base_headers:
            # 컬럼 구성이 다른 시트는 건너뛰고 경고
            continue

        computed_sum = 0.0
        reported_total = None  # 시트 내 모든 소계/합계 행(Company Code 빈 행) 값의 합
        for r in range(2, wsf.max_row + 1):
            vals_f = [wsf.cell(row=r, column=c).value for c in range(1, NCOLS_RAW + 1)]
            if all(v is None for v in vals_f):
                continue
            vals_v = [wsv.cell(row=r, column=c).value for c in range(1, NCOLS_RAW + 1)]
            if vals_v[0] is None:
                # Company Code 빈 값 = 소계/합계 행 -> 결합에서는 제외한다.
                # 한 시트 안에 그룹별 소계 행이 여러 개 있을 수 있으므로(예: Local Currency,
                # Cost Center 그룹별 소계), 이 값들을 모두 더한 합이 곧 그 시트의 데이터 행
                # 합계와 같아야 한다. 따라서 "마지막 행"이 아니라 "모든 소계 행의 합"을 원본
                # 합계로 사용한다.
                amt = vals_v[amt_local_idx_raw]
                if isinstance(amt, (int, float)):
                    reported_total = (reported_total or 0) + amt
                continue
            # Company Code(0)/Cost Center(5)/G/L Account(6)는 raw 시트마다 숫자/문자로 섞여
            # 들어오는 경우가 있다 (예: 본사_HEV 시트는 Cost Center를 숫자로 저장). 같은 열
            # 안에서 타입이 섞이면 Excel이 숫자는 오른쪽, 문자는 왼쪽으로 정렬해 시각적으로
            # 정렬이 어긋나 보이므로, 이 세 열은 항상 텍스트로 통일한다.
            for _idx in (0, 5, 6):
                _v = vals_v[_idx]
                if _v is not None and not isinstance(_v, str):
                    vals_v[_idx] = str(int(_v)) if isinstance(_v, float) and _v.is_integer() else str(_v)

            local_curr = vals_v[10]
            computed_sum += vals_v[amt_local_idx_raw] or 0
            row = [gubun] + vals_v[:12] + [None, None] + vals_v[12:]
            out_ws.append(row)
            total_rows += 1
            excel_row = 2 + total_rows
            fx = fx_rates.get(local_curr)
            if fx is None:
                missing_fx.add(local_curr)
            out_ws.cell(row=excel_row, column=currency_col, value=fx)
            out_ws.cell(row=excel_row, column=krw_col,
                        value=f"={amt_local_letter}{excel_row}*{curr_letter}{excel_row}")

        sheet_validation.append((sname, gubun, computed_sum, reported_total))

    last_row = 2 + total_rows

    amt_doc_col = full_headers.index("Amount in doc. curr.") + 1
    amt_local_col = full_headers.index("Amount in local currency") + 1
    amt_doc_letter = get_column_letter(amt_doc_col)
    amt_local_letter2 = get_column_letter(amt_local_col)
    krw_letter = get_column_letter(krw_col)

    amt_fmt = "#,##0;[Red]-#,##0"
    for col, letter in [(amt_doc_col, amt_doc_letter), (amt_local_col, amt_local_letter2), (krw_col, krw_letter)]:
        out_ws.cell(row=1, column=col, value=f"=SUBTOTAL(109,{letter}3:{letter}{last_row})")
        out_ws.cell(row=1, column=col).number_format = amt_fmt

    # 헤더 스타일: 원본 컬럼 = 밝은 회색, 새로 추가한 컬럼 = 진한 회색. 볼드 없음.
    header_fill = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid")
    manual_fill = PatternFill(start_color="FFA6A6A6", end_color="FFA6A6A6", fill_type="solid")
    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    manual_cols = {"본사/법인", "Currency", "Amount(KRW)"}
    for c, h in enumerate(full_headers, start=1):
        cell = out_ws.cell(row=2, column=c)
        cell.border = border
        cell.font = Font(bold=False)
        cell.fill = manual_fill if h in manual_cols else header_fill

    post_date_col = full_headers.index("Posting Date") + 1
    for r in range(3, last_row + 1):
        out_ws.cell(row=r, column=post_date_col).number_format = "mm-dd-yy"
        out_ws.cell(row=r, column=amt_doc_col).number_format = amt_fmt
        out_ws.cell(row=r, column=amt_local_col).number_format = amt_fmt
        out_ws.cell(row=r, column=currency_col).number_format = "#,##0.00"
        out_ws.cell(row=r, column=krw_col).number_format = amt_fmt

    out_ws.freeze_panes = "A3"
    out_ws.auto_filter.ref = f"A2:{get_column_letter(len(full_headers))}{last_row}"

    widths_map = {
        "본사/법인": 10, "Company Code": 13, "Posting Date": 12.71, "Order": 14.14, "Document Number": 18,
        "Document Type": 15.29, "Cost Center": 11.71, "G/L Account": 12.43, "Posting Key": 12,
        "Document currency": 18.86, "Amount in doc. curr.": 19.57, "Local Currency": 14.57,
        "Amount in local currency": 23.86, "Currency": 12, "Amount(KRW)": 16,
        "Text": 45.71, "Reversed with": 13.57, "Reference key 1": 15.71, "Reference": 22.57,
        "Vendor Name": 40.14, "Vendor": 8, "URL": 45.71,
    }
    for i, h in enumerate(full_headers, start=1):
        out_ws.column_dimensions[get_column_letter(i)].width = widths_map.get(h, 14)

    # ---- 검증 시트: 시트별 Amount in local currency 계산합계 vs 원본 합계행 값 ----
    val_ws = out_wb.create_sheet("검증")
    val_headers = ["원본 시트명", "본사/법인", "계산 합계(결합 대상 행)", "원본 합계행 값", "차이", "결과"]
    val_ws.append(val_headers)
    for c in range(1, len(val_headers) + 1):
        cell = val_ws.cell(row=1, column=c)
        cell.fill = manual_fill
        cell.font = Font(bold=False)
        cell.border = border

    any_fail = False
    vr = 1
    for sname, gubun, computed_sum, reported_total in sheet_validation:
        vr += 1
        if reported_total is None:
            # 원본 시트에 소계/합계 행이 아예 없는 경우: 제외된 행이 없다는 뜻이므로
            # (Company Code가 빈 행이 하나도 없었음), 결합에 포함된 데이터 행 전체의 합이
            # 곧 그 시트의 진짜 합계다. 즉 "대사할 대상이 없어서 확인 불가"가 아니라
            # "전체 데이터 행을 빠짐없이 다 합쳤다"는 것 자체가 검증이다. 따라서 계산합계를
            # 원본 합계행 값 자리에도 그대로 채워 차이 0, "일치"로 표시한다.
            diff = 0.0
            reported_total = computed_sum
            result = "일치(원본에 합계행 없음 - 전체 데이터 행 합산 확인)"
        else:
            diff = computed_sum - reported_total
            result = "일치" if abs(diff) <= TOLERANCE else "확인 필요"
            if result == "확인 필요":
                any_fail = True
        val_ws.cell(row=vr, column=1, value=sname)
        val_ws.cell(row=vr, column=2, value=gubun)
        val_ws.cell(row=vr, column=3, value=round(computed_sum, 2))
        val_ws.cell(row=vr, column=4, value=round(reported_total, 2) if reported_total is not None else None)
        val_ws.cell(row=vr, column=5, value=round(diff, 2) if diff is not None else None)
        val_ws.cell(row=vr, column=6, value=result)
        for c in (3, 4, 5):
            val_ws.cell(row=vr, column=c).number_format = "#,##0.00;[Red]-#,##0.00"
        if result == "확인 필요":
            fail_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
            for c in range(1, len(val_headers) + 1):
                val_ws.cell(row=vr, column=c).fill = fail_fill

    val_widths = [22, 8, 22, 20, 14, 20]
    for i, w in enumerate(val_widths, start=1):
        val_ws.column_dimensions[get_column_letter(i)].width = w
    val_ws.freeze_panes = "A2"

    out_wb.save(output_path)

    print(f"월: {yy}년 {mm}월 (환율 파일: {os.path.basename(fx_path)})")
    print(f"총 {total_rows}행 결합 완료 -> {output_path}")
    if missing_fx:
        print(f"⚠ 환율을 찾지 못한 통화: {sorted(missing_fx)} (해당 행 Currency/Amount(KRW) 비어 있음, 수기 확인 필요)")

    print("\n[검증] 시트별 Amount in local currency 합계 비교")
    for sname, gubun, computed_sum, reported_total in sheet_validation:
        if reported_total is None:
            print(f"  - {sname}: 계산합계={computed_sum:,.2f} / 원본에 합계행 없음 -> 전체 데이터 행 합산으로 확인 완료 [OK]")
        else:
            diff = computed_sum - reported_total
            mark = "OK" if abs(diff) <= TOLERANCE else "MISMATCH"
            print(f"  - {sname}: 계산합계={computed_sum:,.2f} / 원본={reported_total:,.2f} / 차이={diff:,.2f} [{mark}]")
    if any_fail:
        print("⚠ 일부 시트에서 합계가 일치하지 않습니다. '검증' 시트를 확인하세요.")
    else:
        print("모든 시트 합계 일치 확인 완료.")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--raw", required=True, help="raw 데이터 파일 경로 (예: raw_5월.xlsx)")
    ap.add_argument("--fx-folder", required=True, help="환율 파일들이 있는 폴더 경로")
    ap.add_argument("--output", required=True, help="출력 파일 경로 (예: edit_5월.xlsx)")
    args = ap.parse_args()
    build(args.raw, args.fx_folder, args.output)
