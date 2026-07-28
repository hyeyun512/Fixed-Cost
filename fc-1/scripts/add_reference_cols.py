#!/usr/bin/env python3
"""
edit_{월}월.xlsx 에 기준정보_{월}월.xlsx 파일을 참고하여
Company / Cost Center 정보 / G/L Account 정보 열을 추가한다.
맨 앞 본사/법인 구분 열의 헤더명은 '본사/법인'으로 표기한다 (계정구분상의 '구분' 열과 헤더명 중복 방지).
"""
import argparse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side


def norm(v):
    """키 비교용 정규화: 숫자면 int->str, 문자열이면 strip."""
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
        return str(f)
    except ValueError:
        return s


def build(edit_path, ref_path, output_path):
    ref = openpyxl.load_workbook(ref_path, data_only=True)

    # ---------------- 1) company 시트 ----------------
    ws_c = ref['company']
    hdr_hq = [ws_c.cell(row=2, column=c).value for c in (1, 2)]
    hdr_corp = [ws_c.cell(row=7, column=c).value for c in (1, 2)]
    assert hdr_hq == ['Cocd', 'Company'], f"company 시트 본사 헤더 불일치: {hdr_hq}"
    assert hdr_corp[:2] == ['Cocd', 'Company'], f"company 시트 법인 헤더 불일치: {hdr_corp}"

    company_map = {}
    company_map_conflicts = []
    for r in list(range(3, 5)) + list(range(8, ws_c.max_row + 1)):
        cocd = ws_c.cell(row=r, column=1).value
        comp = ws_c.cell(row=r, column=2).value
        if cocd is None or comp is None:
            continue
        k = norm(cocd)
        if k in company_map and company_map[k] != comp:
            company_map_conflicts.append((k, company_map[k], comp, r))
            continue
        company_map.setdefault(k, comp)

    # ---------------- 2) Costcenter 시트 ----------------
    ws_cc = ref['Costcenter']
    hdr_cc = [ws_cc.cell(row=3, column=c).value for c in (3, 4, 7, 8, 9)]
    assert hdr_cc == ['CoCtr', 'Cost Ctr Name', '대조직', '배부조직', '보고용'], f"Costcenter 시트 헤더 불일치: {hdr_cc}"

    costcenter_map = {}
    costcenter_conflicts = []
    for r in range(4, ws_cc.max_row + 1):
        coctr = ws_cc.cell(row=r, column=3).value
        if coctr is None:
            continue
        k = norm(coctr)
        val = (
            ws_cc.cell(row=r, column=4).value,
            ws_cc.cell(row=r, column=7).value,
            ws_cc.cell(row=r, column=8).value,
            ws_cc.cell(row=r, column=9).value,
        )
        if k in costcenter_map and costcenter_map[k] != val:
            costcenter_conflicts.append((k, costcenter_map[k], val, r))
            continue
        costcenter_map.setdefault(k, val)

    # ---------------- 3) 통합계정 시트 ----------------
    ws_ic = ref['통합계정']
    hdr_ic = [ws_ic.cell(row=1, column=c).value for c in (1, 2, 3, 4)]
    assert hdr_ic == ['G/L', '소계정', '대계정', '분류'], f"통합계정 시트 헤더 불일치: {hdr_ic}"

    account_map = {}
    account_conflicts = []
    for r in range(2, ws_ic.max_row + 1):
        gl = ws_ic.cell(row=r, column=1).value
        if gl is None:
            continue
        k = norm(gl)
        val = (
            ws_ic.cell(row=r, column=2).value,
            ws_ic.cell(row=r, column=3).value,
            ws_ic.cell(row=r, column=4).value,
        )
        if k in account_map and account_map[k] != val:
            account_conflicts.append((k, account_map[k], val, r))
            continue
        account_map.setdefault(k, val)

    # ---------------- 4) 계정구분 시트: [구분] 테이블 ----------------
    ws_gb = ref['계정구분']
    hdr_gubun = [ws_gb.cell(row=3, column=c).value for c in (2, 3)]
    assert hdr_gubun == ['대계정(re)', '구분'], f"계정구분 시트 [구분] 헤더 불일치: {hdr_gubun}"

    gubun_map = {}
    r = 4
    while True:
        key = ws_gb.cell(row=r, column=2).value
        val = ws_gb.cell(row=r, column=3).value
        if key is None:
            break
        if str(key).startswith('(*)'):
            break
        gubun_map[key] = val
        r += 1

    # ---------------- 5) 계정구분 시트: [4대보험료] 테이블 ----------------
    hdr_4bh = [ws_gb.cell(row=3, column=c).value for c in (6, 7, 8, 9, 10)]
    assert hdr_4bh == ['Company', 'G/L Account', '소계정', '대계정(원본)', '대계정(re)'], f"계정구분 시트 [4대보험료] 헤더 불일치: {hdr_4bh}"

    fourins_all_gl = set()
    fourins_specific = set()
    r = 4
    while True:
        comp = ws_gb.cell(row=r, column=6).value
        gl = ws_gb.cell(row=r, column=7).value
        target = ws_gb.cell(row=r, column=10).value
        if comp is None or gl is None:
            break
        if comp == 'All':
            fourins_all_gl.add(norm(gl))
        else:
            fourins_specific.add((comp, norm(gl)))
        assert target == '53 4대보험료'
        r += 1

    # ================= edit 파일 처리 =================
    wb = openpyxl.load_workbook(edit_path, data_only=False)
    ws = wb['edit']
    old_headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    last_row = ws.max_row

    # 맨 앞 열은 fc-1 스킬이 만든 '구분'(본사/법인 구분자) 열이었으나, 이제 헤더명을 '본사/법인'으로 표기한다.
    new_headers = (
        ['본사/법인', 'Company Code', 'Company', 'Posting Date', 'Order', 'Document Number', 'Document Type',
         'Cost Center', 'Cost Ctr Name', '대조직', '배부조직', '보고용', '보고용(re)',
         'G/L Account', '소계정', '대계정', '대계정(re)', '구분', '분류',
         'Posting Key', 'Document currency', 'Amount in doc. curr.', 'Local Currency',
         'Amount in local currency', 'Currency', 'Amount(KRW)',
         'Text', 'Reversed with', 'Reference key 1', 'Reference', 'Vendor Name', 'Vendor', 'URL']
    )
    new_col = {h: i + 1 for i, h in enumerate(new_headers)}
    manual_cols = {'본사/법인', 'Currency', 'Amount(KRW)',
                   'Company', 'Cost Ctr Name', '대조직', '배부조직', '보고용', '보고용(re)',
                   '소계정', '대계정', '대계정(re)', '구분', '분류'}

    # old 파일의 맨 앞 '구분'(본사/법인) 열은 첫 번째 등장 위치 기준으로 읽는다
    # (이미 한 번 enrichment 된 파일을 다시 입력해도 안전하도록).
    old_idx_first = {}
    for i, h in enumerate(old_headers):
        old_idx_first.setdefault(h, i + 1)
    old_data = {}
    for old_h, old_c in old_idx_first.items():
        col_vals = []
        for r in range(3, last_row + 1):
            cell = ws.cell(row=r, column=old_c)
            col_vals.append(cell.value)
        old_data[old_h] = col_vals
    old_number_formats = {}
    for old_h, old_c in old_idx_first.items():
        old_number_formats[old_h] = ws.cell(row=3, column=old_c).number_format

    # 맨 앞 본사/법인 열 원본 데이터: 이전 실행 결과라면 헤더가 이미 '본사/법인'일 수도, fc-1 원본이라면 '구분'일 수도 있음
    hq_corp_key = '본사/법인' if '본사/법인' in old_data else '구분'

    missing_company = set()
    missing_costcenter = set()
    missing_gl = set()
    missing_gubun = set()

    HQ_CODES = {'1000', 'E100'}
    OVERSEAS_DEFAULT_BOGOYONG = '7. 해외법인(법인)'
    BOGOYONG_RE_TRIGGERS = {'6.해외법인(주재원)', OVERSEAS_DEFAULT_BOGOYONG}

    n = last_row - 2
    company_col_vals = []
    costctr_name_vals, daejojik_vals, baebujojik_vals, bogoyong_vals, bogoyong_re_vals = [], [], [], [], []
    sojaejong_vals, daegyejong_vals, daegyejong_re_vals, gubun_vals, bunryu_vals = [], [], [], [], []

    for i in range(n):
        cc_raw = old_data['Company Code'][i]
        cost_raw = old_data['Cost Center'][i]
        gl_raw = old_data['G/L Account'][i]

        cc_k = norm(cc_raw)
        cost_k = norm(cost_raw)
        gl_k = norm(gl_raw)

        company = company_map.get(cc_k)
        if company is None and cc_k is not None:
            missing_company.add(cc_raw)

        cc_vals = costcenter_map.get(cost_k)
        if cc_vals is None:
            if cost_k is not None:
                missing_costcenter.add(cost_raw)
            if cc_k not in HQ_CODES:
                # 해외법인(법인) 기본값: 대조직/배부조직도 Company 열 값과 동일하게 기재
                cc_vals = (None, company, company, OVERSEAS_DEFAULT_BOGOYONG)
            else:
                cc_vals = (None, None, None, None)
        costctr_name, daejojik, baebujojik, bogoyong = cc_vals

        # 주재원 보정: 배부조직 값에 '주재원'이 포함된 경우, Company는 대조직 값과 동일하게 설정
        if baebujojik is not None and '주재원' in str(baebujojik):
            company = daejojik

        company_col_vals.append(company)
        costctr_name_vals.append(costctr_name)
        daejojik_vals.append(daejojik)
        baebujojik_vals.append(baebujojik)
        bogoyong_vals.append(bogoyong)

        if bogoyong in BOGOYONG_RE_TRIGGERS:
            bogoyong_re_vals.append(company)
        else:
            bogoyong_re_vals.append(bogoyong)

        acc_vals = account_map.get(gl_k)
        if acc_vals is None:
            if gl_k is not None:
                missing_gl.add(gl_raw)
            acc_vals = (None, None, None)
        sojaejong, daegyejong, bunryu = acc_vals
        sojaejong_vals.append(sojaejong)
        daegyejong_vals.append(daegyejong)

        daegyejong_re = daegyejong
        if gl_k in fourins_all_gl:
            daegyejong_re = '53 4대보험료'
        elif company is not None and (company, gl_k) in fourins_specific:
            daegyejong_re = '53 4대보험료'
        daegyejong_re_vals.append(daegyejong_re)

        gubun_val = gubun_map.get(daegyejong_re)
        if gubun_val is None and daegyejong_re is not None:
            missing_gubun.add(daegyejong_re)
        gubun_vals.append(gubun_val)
        bunryu_vals.append(bunryu)

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = 'edit'
    if '검증' in wb.sheetnames:
        src_val = wb['검증']
        tgt_val = out_wb.create_sheet('검증')
        for row in src_val.iter_rows():
            for cell in row:
                if cell.value is not None:
                    tgt_val.cell(row=cell.row, column=cell.column, value=cell.value)
                tgt_val.cell(row=cell.row, column=cell.column).number_format = cell.number_format
        for i, col_dim in src_val.column_dimensions.items():
            tgt_val.column_dimensions[i].width = col_dim.width
        tgt_val.freeze_panes = src_val.freeze_panes

    out_ws.append([None] * len(new_headers))
    out_ws.append(new_headers)

    col_letter = {h: get_column_letter(new_col[h]) for h in new_headers}

    for i in range(n):
        row_vals = [None] * len(new_headers)
        row_vals[new_col['본사/법인'] - 1] = old_data[hq_corp_key][i]
        row_vals[new_col['Company Code'] - 1] = old_data['Company Code'][i]
        row_vals[new_col['Company'] - 1] = company_col_vals[i]
        row_vals[new_col['Posting Date'] - 1] = old_data['Posting Date'][i]
        row_vals[new_col['Order'] - 1] = old_data['Order'][i]
        row_vals[new_col['Document Number'] - 1] = old_data['Document Number'][i]
        row_vals[new_col['Document Type'] - 1] = old_data['Document Type'][i]
        row_vals[new_col['Cost Center'] - 1] = old_data['Cost Center'][i]
        row_vals[new_col['Cost Ctr Name'] - 1] = costctr_name_vals[i]
        row_vals[new_col['대조직'] - 1] = daejojik_vals[i]
        row_vals[new_col['배부조직'] - 1] = baebujojik_vals[i]
        row_vals[new_col['보고용'] - 1] = bogoyong_vals[i]
        row_vals[new_col['보고용(re)'] - 1] = bogoyong_re_vals[i]
        row_vals[new_col['G/L Account'] - 1] = old_data['G/L Account'][i]
        row_vals[new_col['소계정'] - 1] = sojaejong_vals[i]
        row_vals[new_col['대계정'] - 1] = daegyejong_vals[i]
        row_vals[new_col['대계정(re)'] - 1] = daegyejong_re_vals[i]
        row_vals[new_col['구분'] - 1] = gubun_vals[i]
        row_vals[new_col['분류'] - 1] = bunryu_vals[i]

        out_ws.append(row_vals)
        rr = out_ws.max_row

        out_ws.cell(row=rr, column=new_col['Posting Key'], value=old_data['Posting Key'][i])
        out_ws.cell(row=rr, column=new_col['Document currency'], value=old_data['Document currency'][i])
        out_ws.cell(row=rr, column=new_col['Amount in doc. curr.'], value=old_data['Amount in doc. curr.'][i])
        out_ws.cell(row=rr, column=new_col['Local Currency'], value=old_data['Local Currency'][i])
        out_ws.cell(row=rr, column=new_col['Amount in local currency'], value=old_data['Amount in local currency'][i])
        out_ws.cell(row=rr, column=new_col['Currency'], value=old_data['Currency'][i])
        amt_local_letter = col_letter['Amount in local currency']
        curr_letter = col_letter['Currency']
        out_ws.cell(row=rr, column=new_col['Amount(KRW)'],
                    value=f"={amt_local_letter}{rr}*{curr_letter}{rr}")
        out_ws.cell(row=rr, column=new_col['Text'], value=old_data['Text'][i])
        out_ws.cell(row=rr, column=new_col['Reversed with'], value=old_data['Reversed with'][i])
        out_ws.cell(row=rr, column=new_col['Reference key 1'], value=old_data['Reference key 1'][i])
        out_ws.cell(row=rr, column=new_col['Reference'], value=old_data['Reference'][i])
        out_ws.cell(row=rr, column=new_col['Vendor Name'], value=old_data['Vendor Name'][i])
        out_ws.cell(row=rr, column=new_col['Vendor'], value=old_data['Vendor'][i])
        out_ws.cell(row=rr, column=new_col['URL'], value=old_data['URL'][i])

    new_last_row = 2 + n

    amt_fmt = old_number_formats.get('Amount in doc. curr.', '#,##0;[Red]-#,##0')
    for h in ['Amount in doc. curr.', 'Amount in local currency', 'Amount(KRW)']:
        letter = col_letter[h]
        out_ws.cell(row=1, column=new_col[h], value=f"=SUBTOTAL(109,{letter}3:{letter}{new_last_row})")
        out_ws.cell(row=1, column=new_col[h]).number_format = amt_fmt

    header_fill = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid")
    manual_fill = PatternFill(start_color="FFA6A6A6", end_color="FFA6A6A6", fill_type="solid")
    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, h in enumerate(new_headers, start=1):
        cell = out_ws.cell(row=2, column=c)
        cell.border = border
        cell.font = Font(bold=False)
        cell.fill = manual_fill if h in manual_cols else header_fill

    post_date_letter_col = new_col['Posting Date']
    for r in range(3, new_last_row + 1):
        out_ws.cell(row=r, column=post_date_letter_col).number_format = 'mm-dd-yy'
        for h in ['Amount in doc. curr.', 'Amount in local currency', 'Amount(KRW)']:
            out_ws.cell(row=r, column=new_col[h]).number_format = amt_fmt
        out_ws.cell(row=r, column=new_col['Currency']).number_format = old_number_formats.get('Currency', '#,##0.00')

    out_ws.freeze_panes = 'A3'
    out_ws.auto_filter.ref = f"A2:{get_column_letter(len(new_headers))}{new_last_row}"

    widths_map = {
        '본사/법인': 10, 'Company Code': 13, 'Company': 10, 'Posting Date': 12.71, 'Order': 14.14,
        'Document Number': 18, 'Document Type': 15.29, 'Cost Center': 11.71, 'Cost Ctr Name': 16,
        '대조직': 12, '배부조직': 12, '보고용': 14, '보고용(re)': 14,
        'G/L Account': 12.43, '소계정': 22, '대계정': 14,
        '대계정(re)': 14, '구분': 10, '분류': 10, 'Posting Key': 12,
        'Document currency': 18.86, 'Amount in doc. curr.': 19.57, 'Local Currency': 14.57,
        'Amount in local currency': 23.86, 'Currency': 12, 'Amount(KRW)': 16,
        'Text': 45.71, 'Reversed with': 13.57, 'Reference key 1': 15.71, 'Reference': 22.57,
        'Vendor Name': 40.14, 'Vendor': 8, 'URL': 45.71,
    }
    for i, h in enumerate(new_headers, start=1):
        out_ws.column_dimensions[get_column_letter(i)].width = widths_map.get(h, 14)

    out_wb.save(output_path)

    print(f"완료: {output_path}")
    print(f"총 데이터 행: {n}")
    print(f"\n[매핑 확인]")
    print(f"company_map 충돌: {company_map_conflicts}")
    print(f"costcenter_map 충돌: {costcenter_conflicts}")
    print(f"account_map 충돌: {account_conflicts}")
    print(f"\n미매칭 Company Code: {sorted(missing_company, key=str)}")
    print(f"미매칭 Cost Center: {sorted(missing_costcenter, key=str)}")
    print(f"미매칭 G/L Account: {sorted(missing_gl, key=str)}")
    print(f"미매칭 대계정(re) (구분 테이블에 없음): {sorted(missing_gubun, key=str)}")

    from collections import Counter
    bogoyong_counter = Counter(bogoyong_vals)
    print(f"\n보고용 = '7. 해외법인(법인)' 건수: {bogoyong_counter.get(OVERSEAS_DEFAULT_BOGOYONG, 0)}")
    bogoyong_re_overridden = sum(1 for b in bogoyong_vals if b in BOGOYONG_RE_TRIGGERS)
    print(f"보고용(re)가 Company 값으로 대체된 건수: {bogoyong_re_overridden}")


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--edit', required=True)
    ap.add_argument('--ref', required=True)
    ap.add_argument('--output', required=True)
    args = ap.parse_args()
    build(args.edit, args.ref, args.output)
