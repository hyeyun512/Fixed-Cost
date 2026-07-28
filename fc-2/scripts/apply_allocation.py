#!/usr/bin/env python3
"""
fc-2: edit_{월}월.xlsx 파일에 Supabase 리소스배부율을 반영해
edit_{월}월(배부ver).xlsx 파일을 생성한다.

배부율(%) 열은 값이 아니라 워크북에 함께 추가되는 '배부율(Supabase)' 참고 시트를
VLOOKUP으로 참조하는 수식으로 채운다. 배부기준이 배부율표에 없어 매칭 실패하면
(예: '별도배부율') IFERROR로 0(0%)을 반환한다 — 배부후금액 수식이 깨지지 않도록
항상 숫자를 반환해야 하며, 텍스트("확인요망" 등)를 반환해서는 안 된다.

사용법:
    python apply_allocation.py <edit파일 경로> <supabase_rates.json> <output.xlsx>
"""
import json
import sys

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

DIVISIONS = [
    ("stb", "STB"),
    ("mobility", "Mobility"),
    ("evcs_domestic", "EVCS(국내)"),
    ("evcs_overseas", "EVCS(해외)"),
    ("humax_common", "Humax(공통)"),
    ("building", "건물"),
    ("h_mobility", "H.Mobility"),
    ("h_ev", "H.EV"),
    ("hiparking", "하이파킹"),
    ("peoplecar", "피플카"),
    ("winercom", "위너콤"),
    ("holdings", "홀딩스"),
    ("h_networks", "H.Networks"),
]
N = len(DIVISIONS)

RATE_FILL = "E8F0FE"
AMT_FILL = "FFF6E0"
COND_FILL = "FDE9D9"
GRID_BORDER_COLOR = "D9D9D9"
EXCLUDED_BASIS_LABEL = "실적 제외"


def find_basis_column(ws, header_row=2):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == "배부기준":
            return col
    raise ValueError("'배부기준' 헤더를 찾지 못했습니다.")


def find_amount_krw_column(ws, header_row=2):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == "Amount(KRW)":
            return col
    raise ValueError("'Amount(KRW)' 헤더를 찾지 못했습니다.")


def build_rates_dict(supabase_rows):
    rates = {}
    for r in supabase_rows:
        basis = r.get("basis")
        if not basis:
            continue
        vals = [float(r.get(key) or 0.0) for key, _ in DIVISIONS]
        rates[basis] = vals
    return rates


def main():
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)
    edit_path, supabase_json_path, output_path = sys.argv[1:4]

    with open(supabase_json_path, encoding="utf-8") as f:
        supabase_rows = json.load(f)
    rates = build_rates_dict(supabase_rows)

    wb = openpyxl.load_workbook(edit_path)
    ws = wb["edit"]
    MAXROW = ws.max_row
    z_col = find_amount_krw_column(ws)
    z_letter = get_column_letter(z_col)
    ai_col = find_basis_column(ws)
    ai_letter = get_column_letter(ai_col)

    # ---- 배부기준 값 중 Supabase 배부율표에 없는 key 찾기 (LUT에는 넣지 않는다 -> VLOOKUP 실패 -> IFERROR로 0%) ----
    used_basis = set()
    for row in range(3, MAXROW + 1):
        v = ws.cell(row=row, column=ai_col).value
        if v is not None:
            used_basis.add(v)
    unmatched = sorted(b for b in used_basis if b not in rates and b != EXCLUDED_BASIS_LABEL)
    excluded_present = EXCLUDED_BASIS_LABEL in used_basis
    if EXCLUDED_BASIS_LABEL not in rates:
        rates[EXCLUDED_BASIS_LABEL] = [0.0] * N
    # 매칭 안 되는 key는 절대 rates(=LUT)에 추가하지 않는다. VLOOKUP이 실패해야
    # IFERROR가 0(0%)을 반환한다.

    # ---- 참고용 배부율표 시트 ----
    LUT_NAME = "배부율(Supabase)"
    if LUT_NAME in wb.sheetnames:
        del wb[LUT_NAME]
    lut = wb.create_sheet(LUT_NAME)
    lut.append(["배부기준"] + [label for _, label in DIVISIONS] + ["TOTAL"])
    for c in lut[1]:
        c.font = Font(bold=True)
    for basis, vals in rates.items():
        lut.append([basis] + vals + [sum(vals)])
    lut.freeze_panes = "A2"
    for row in lut.iter_rows(min_row=2, max_row=lut.max_row, min_col=2, max_col=2 + N):
        for cell in row:
            cell.number_format = "0.00%"
    lut.column_dimensions["A"].width = 26
    lut_max_row = lut.max_row
    lut_last_col_letter = get_column_letter(2 + N)  # A(배부기준) + N개 division + TOTAL

    # ---- 열 배치 ----
    RATE_START = ai_col + 1
    RATE_SUM_COL = RATE_START + N
    AMT_START = RATE_SUM_COL + 1
    TOTAL_COL = AMT_START + N

    header_font = Font(bold=True)
    rate_fill = PatternFill("solid", fgColor=Color(rgb="FF" + RATE_FILL))
    amt_fill = PatternFill("solid", fgColor=Color(rgb="FF" + AMT_FILL))
    z_fmt = ws.cell(row=1, column=z_col).number_format
    if not z_fmt or z_fmt == "General":
        z_fmt = "#,##0;[Red]-#,##0"

    for i, (_, label) in enumerate(DIVISIONS):
        c = ws.cell(row=2, column=RATE_START + i, value=f"{label}(%)")
        c.font = header_font
        c.fill = rate_fill
    c = ws.cell(row=2, column=RATE_SUM_COL, value="합계(%)")
    c.font = header_font
    c.fill = rate_fill
    for i, (_, label) in enumerate(DIVISIONS):
        c = ws.cell(row=2, column=AMT_START + i, value=f"{label}(KRW)")
        c.font = header_font
        c.fill = amt_fill
    c = ws.cell(row=2, column=TOTAL_COL, value="합계(KRW)")
    c.font = header_font
    c.fill = amt_fill

    # row1: 배부후금액/합계금액 열은 기존 관례대로 SUBTOTAL(필터 연동), Z열과 동일한 숫자 서식
    for i in range(N):
        col = AMT_START + i
        letter = get_column_letter(col)
        cell = ws.cell(row=1, column=col, value=f"=SUBTOTAL(109,{letter}3:{letter}{MAXROW})")
        cell.number_format = z_fmt
    tot_letter = get_column_letter(TOTAL_COL)
    cell = ws.cell(row=1, column=TOTAL_COL, value=f"=SUBTOTAL(109,{tot_letter}3:{tot_letter}{MAXROW})")
    cell.number_format = z_fmt

    # data rows
    for row in range(3, MAXROW + 1):
        r1 = get_column_letter(RATE_START)
        r2 = get_column_letter(RATE_START + N - 1)
        for i in range(N):
            col = RATE_START + i
            formula = (
                f'=IFERROR(VLOOKUP(${ai_letter}{row},'
                f"'{LUT_NAME}'!$A$2:${lut_last_col_letter}${lut_max_row},{i + 2},FALSE),0)"
            )
            cell = ws.cell(row=row, column=col, value=formula)
            cell.number_format = "0.00%"
        cell = ws.cell(row=row, column=RATE_SUM_COL, value=f"=SUM({r1}{row}:{r2}{row})")
        cell.number_format = "0.00%"
        for i in range(N):
            rate_letter = get_column_letter(RATE_START + i)
            col = AMT_START + i
            cell = ws.cell(
                row=row, column=col,
                value=f"=${z_letter}{row}*{rate_letter}{row}",
            )
            cell.number_format = z_fmt
        a1 = get_column_letter(AMT_START)
        a2 = get_column_letter(AMT_START + N - 1)
        cell = ws.cell(row=row, column=TOTAL_COL, value=f"=SUM({a1}{row}:{a2}{row})")
        cell.number_format = z_fmt

    # ---- 합계(%) 열 조건부 서식 ----
    rs_letter = get_column_letter(RATE_SUM_COL)
    cond_fill = PatternFill(patternType="solid",
                             fgColor=Color(rgb="FF" + COND_FILL),
                             bgColor=Color(rgb="FF" + COND_FILL))
    formula = (f'AND(ABS({rs_letter}3-1)>0.001,'
               f'NOT(AND({rs_letter}3=0,${ai_letter}3="{EXCLUDED_BASIS_LABEL}")))')
    ws.conditional_formatting.add(
        f"{rs_letter}3:{rs_letter}{MAXROW}",
        FormulaRule(formula=[formula], fill=cond_fill, stopIfTrue=False),
    )

    thin = Side(style="thin", color=GRID_BORDER_COLOR)
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(2, MAXROW + 1):
        ws.cell(row=row, column=RATE_SUM_COL).border = grid_border

    # ---- 열 너비 ----
    for i in range(N):
        ws.column_dimensions[get_column_letter(RATE_START + i)].width = 11
    ws.column_dimensions[rs_letter].width = 11
    for i in range(N):
        ws.column_dimensions[get_column_letter(AMT_START + i)].width = 15
    ws.column_dimensions[get_column_letter(TOTAL_COL)].width = 15

    # ---- 자동 필터를 새 열까지 확장 ----
    total_letter = get_column_letter(TOTAL_COL)
    ws.auto_filter.ref = f"A2:{total_letter}{MAXROW}"

    # ---- 검증 시트에 요약 추가 ----
    if "검증" in wb.sheetnames:
        vws = wb["검증"]
    else:
        vws = wb.create_sheet("검증")
    start_row = vws.max_row + 2 if vws.max_row > 1 else 1
    vws.cell(row=start_row, column=1,
              value="배부 검증 (Supabase 리소스배부율 반영, 수식/VLOOKUP 버전)").font = Font(bold=True)
    hdr = ["구분", "건수", "비고"]
    for j, h in enumerate(hdr):
        vws.cell(row=start_row + 1, column=1 + j, value=h).font = Font(bold=True)
    matched = sum(1 for b in used_basis if b in rates and b not in unmatched and b != EXCLUDED_BASIS_LABEL)
    r = start_row + 2
    rows_summary = [
        ["배부율표 매칭", matched, "AI열 배부기준 값이 '배부율(Supabase)' 시트와 VLOOKUP으로 매칭되어 배부율 합계 100%, Amount(KRW)와 합계(KRW) 일치"],
    ]
    if excluded_present:
        rows_summary.append([EXCLUDED_BASIS_LABEL, "-", "배부율표에서 TOTAL=0으로 정의된 항목 (설계상 배부율 0%, 합계(%)열 채우기 없음)"])
    if unmatched:
        rows_summary.append(["배부율표 미매칭 (0% 적용)", len(unmatched),
                              "배부율표에 없는 배부기준: " + ", ".join(unmatched) +
                              " → VLOOKUP 매칭 실패시 IFERROR로 0% 적용, 건별 확인 필요"])
    for d in rows_summary:
        vws.cell(row=r, column=1, value=d[0])
        vws.cell(row=r, column=2, value=d[1])
        vws.cell(row=r, column=3, value=d[2])
        r += 1
    vws.column_dimensions["A"].width = 42
    vws.column_dimensions["C"].width = 90

    wb.save(output_path)
    print(f"완료: {output_path}")
    print(f"미매칭 배부기준 (0% 처리): {unmatched if unmatched else '없음'}")


if __name__ == "__main__":
    main()
