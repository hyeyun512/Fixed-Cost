#!/usr/bin/env python3
"""
fc-2 결과 파일(edit_{월}월(배부ver).xlsx)의 배부 합계를 검증한다.

사용법:
    python verify.py <edit_{월}월(배부ver).xlsx>

LibreOffice로 미리 재계산(headless convert-to xlsx)한 파일을 넣어야
수식 값이 캐시되어 있어 정확히 읽힌다. 재계산을 안 한 파일을 넣으면
캐시된 값이 없어 전부 None으로 나올 수 있다.

배부기준별로 그룹핑해서 Amount(KRW) 합계와 합계(KRW)(배부후금액 합계)를
비교하고, 차이가 나는 배부기준을 리포트한다. "실적 제외"는 설계상 배부율
0%이므로 차이가 나는 게 정상이다 — 그 외 배부기준에서 차이가 나면 문제가
있는 것이다.
"""
import sys
from collections import defaultdict

import openpyxl


def find_col(ws, header, header_row=2):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == header:
            return col
    raise ValueError(f"'{header}' 헤더를 찾지 못했습니다.")


def main():
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(1)
    path = sys.argv[1]
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["edit"]

    ai_col = find_col(ws, "배부기준")
    z_col = find_col(ws, "Amount(KRW)")
    total_col = find_col(ws, "합계(KRW)")

    sums_z = defaultdict(float)
    sums_total = defaultdict(float)
    counts = defaultdict(int)

    for row in range(3, ws.max_row + 1):
        basis = ws.cell(row=row, column=ai_col).value
        z = ws.cell(row=row, column=z_col).value or 0
        total = ws.cell(row=row, column=total_col).value or 0
        sums_z[basis] += z
        sums_total[basis] += total
        counts[basis] += 1

    print(f"{'배부기준':30s} {'건수':>6s} {'Amount(KRW)':>16s} {'합계(KRW)':>16s} {'차이':>16s}")
    problem = []
    for basis in sorted(sums_z, key=lambda b: -abs(sums_z[b] - sums_total[b])):
        diff = round(sums_z[basis] - sums_total[basis], 2)
        flag = ""
        if abs(diff) > 1 and basis != "실적 제외":
            flag = "  <-- 확인 필요"
            problem.append(basis)
        print(f"{str(basis):30s} {counts[basis]:6d} {sums_z[basis]:16,.0f} {sums_total[basis]:16,.0f} {diff:16,.0f}{flag}")

    print()
    if problem:
        print(f"※ '실적 제외' 외에 차이가 발생한 배부기준: {problem}")
        print("  -> 배부율표 반영 또는 수식에 문제가 있을 수 있으니 확인이 필요합니다.")
    else:
        print("※ '실적 제외'(설계상 0%) 외에는 모두 Amount(KRW)와 합계(KRW)가 일치합니다.")


if __name__ == "__main__":
    main()
